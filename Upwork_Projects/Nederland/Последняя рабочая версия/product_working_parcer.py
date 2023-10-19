import re
import openpyxl
import requests
from bs4 import BeautifulSoup

# Открываем файл Excel
workbook = openpyxl.load_workbook('/Users/alexey_muzgin/Upwork/Cities.xlsx')
sheet = workbook.active

# Проходим по всем ячейкам в первом столбце (столбец A)
for idx, row in enumerate(sheet.iter_rows(min_col=1, max_col=1, values_only=True), start=1):
    city = row[0].lower()  # Преобразуем в прописные буквы

    # Инициализируем переменную dop
    dop = None

    # Проверяем, есть ли текст в скобках в переменной city
    match = re.search(r'\((.*?)\)', city)
    if match:
        dop = match.group(1)
        # Убираем текст в скобках из переменной city
        city = re.sub(r'\(.*?\)', '', city).strip()

    # Заменяем пробелы на дефисы, если city состоит из двух слов
    if " " in city:
        city = city.replace(" ", "-")

    # Отправляем запрос на веб-сайт
    url = f'https://allecijfers.nl/woonplaats/{city}'
    response = requests.get(url)
    print(response)
    soup = BeautifulSoup(response.content, 'html.parser')

    # Парсим страницу
    inwoners_tag = soup.find('td', text='Inwoners')
    if inwoners_tag:
        aantal_tag = inwoners_tag.find_next('td')
        if aantal_tag:
            inw = aantal_tag.text.replace('.', '')

            # Проверяем, нужно ли добавить текст из скобок в переменную dop
            if dop:
                dop = f'({dop})'
            else:
                dop = ''

            # Записываем результат во второй столбец файла
            sheet.cell(row=idx, column=2, value=f'{inw}')

# Сохраняем изменения в файле
workbook.save('/Users/alexey_muzgin/Upwork/Cities.xlsx')
