import openpyxl
import re
import requests
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.firefox.service import Service

# Указываем путь к веб-драйверу Firefox
firefox_driver_path = '/opt/homebrew/bin/geckodriver'
service = Service(firefox_driver_path)
driver = webdriver.Firefox(service=service)

driver.get('https://www.google.com/')

# Открываем файл Excel
workbook = openpyxl.load_workbook('/Users/alexey_muzgin/Upwork/Cities.xlsx')
sheet = workbook.active

# Проходим по всем ячейкам в первом столбце (столбец A)
for idx, row in enumerate(sheet.iter_rows(min_col=1, max_col=1, values_only=True), start=1):
    city = row[0].lower()  # Преобразуем в прописные буквы
    print(city)

    # Инициализируем переменную dop
    dop = ''

    # Проверяем, есть ли текст в скобках в переменной city
    match = re.search(r'\((.*?)\)', city)
    if match:
        dop = match.group(1)
        # Убираем текст в скобках из переменной city
        city = re.sub(r'\(.*?\)', '', city).strip()

    # Заменяем пробелы на дефисы, если city состоит из двух слов
    if " " in city:
        city = city.replace(" ", "-")

    # Отправляем запрос на веб-сайт
    url = f'https://allecijfers.nl/woonplaats/{city}'
    response = requests.get(url)

    # Если запрос не удался, выполняем поиск на Google и находим ссылку на allecijfers.nl
    if response.status_code != 200:
        print('SELENIUM')
        search_query = f'{city}+{dop}+inwoners'
        google_url = f'https://www.google.com/search?q={search_query}'
        driver.get(google_url)
        google_soup = BeautifulSoup(driver.page_source, 'html.parser')
        
        # Ищем ссылку на allecijfers.nl
        allecijfers_link = google_soup.find('a', href=re.compile(r'https://allecijfers.nl/woonplaats/.+'))
        if allecijfers_link:
            print(allecijfers_link)
            city = re.search(r'https://allecijfers.nl/woonplaats/(.+)', allecijfers_link['href']).group(1)
            url = f'https://allecijfers.nl/woonplaats/{city}'
            response = requests.get(url)
        else:
            # Попробуем другие варианты ссылок
            urls_to_try = [
                f'https://allecijfers.nl/wijk/{city}',
                f'https://allecijfers.nl/gemeente/{city}',
                f'https://allecijfers.nl/buurt/{city}'
            ]
            for alternative_url in urls_to_try:
                response = requests.get(alternative_url)
                if response.status_code == 200:
                    url = alternative_url
                    break

    if response.status_code == 200:
        # Парсим страницу
        soup = BeautifulSoup(response.content, 'html.parser')
        inwoners_tag = soup.find('td', text='Inwoners')
        if inwoners_tag:
            aantal_tag = inwoners_tag.find_next('td')
            if aantal_tag:
                inw = aantal_tag.text.replace('.', '')
                print(inw)

                # Записываем результат во второй столбец файла
                sheet.cell(row=idx, column=2, value=f'{inw}')
            else:
                print(f'Не удалось получить данные для города: {city}')
        else:
            print(f'Не удалось получить данные для города: {city}')
    else:
        print(f'Не удалось получить данные для города: {city}')


# Сохраняем изменения в файле
workbook.save('/Users/alexey_muzgin/Upwork/Cities.xlsx')
